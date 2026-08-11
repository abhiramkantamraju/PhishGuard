"""
Measures PhishGuard against a labelled email dataset.

This is the evidence artifact for the project's accuracy claims: point it at a
CSV or XLSX with `label` and `email_text` columns and it reports the full
confusion matrix plus accuracy, precision, recall and F1.

    python evaluate_dataset.py sample_data/sample_emails.csv
    python evaluate_dataset.py data/realistic_phishing/phishing_focused_holdout.csv --quiet
    python evaluate_dataset.py data/realistic_phishing/realistic_emails.csv --json results.json

Only the offline rule engine is exercised. The Google Safe Browsing and WHOIS
checks are deliberately excluded: they need live network access, which would
make this run slow, rate-limited and non-reproducible. What is measured here is
exactly what runs with no configuration and no internet connection.
"""

import argparse
import csv
import json
import sys
from pathlib import Path

from phishguard import DANGEROUS, SUSPICIOUS, analyze

# Real-world email bodies can exceed the csv module's default 128KB field size
# limit; raise it (capped to what the platform's C long supports).
csv.field_size_limit(min(sys.maxsize, 2**31 - 1))

PHISHING_LEVELS = {SUSPICIOUS, DANGEROUS}

PHISHING_LABELS = {"phishing", "phish", "malicious", "spam", "1", "true"}
LEGITIMATE_LABELS = {"legitimate", "legit", "safe", "ham", "0", "false"}


def normalize_label(label):
    value = str(label).strip().lower()
    if value in PHISHING_LABELS:
        return "phishing"
    if value in LEGITIMATE_LABELS:
        return "legitimate"
    raise ValueError(f"Unsupported label: {label!r}")


def predict_label(email_text):
    """`(prediction, risk_level, score, flags)` for one email."""
    analysis = analyze(email_text)
    prediction = "phishing" if analysis.risk_level in PHISHING_LEVELS else "legitimate"
    return prediction, analysis.risk_level, analysis.score, analysis.flags


def read_rows(path):
    path = Path(path)
    required_columns = {"label", "email_text"}

    if path.suffix.lower() in {".csv", ".tsv"}:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        with open(path, newline="", encoding="utf-8", errors="replace") as csv_file:
            reader = csv.DictReader(csv_file, delimiter=delimiter)
            missing = required_columns - set(reader.fieldnames or [])
            if missing:
                raise ValueError(
                    f"{path.name} is missing required columns: {', '.join(sorted(missing))}"
                )
            yield from reader
        return

    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value else "" for value in next(rows)]
        missing = required_columns - set(headers)
        if missing:
            raise ValueError(
                f"{path.name} is missing required columns: {', '.join(sorted(missing))}"
            )
        for row in rows:
            # strict=False on purpose: openpyxl trims trailing empty cells, so a
            # row can legitimately be shorter than the header list.
            yield dict(zip(headers, row, strict=False))
        return

    raise ValueError("Dataset must be a .csv, .tsv or .xlsx file.")


def evaluate_dataset(path, show_rows=True):
    counts = {"tp": 0, "fp": 0, "tn": 0, "fn": 0}
    level_counts = {}
    total = 0

    for row in read_rows(path):
        actual = normalize_label(row["label"])
        email_text = "" if row["email_text"] is None else str(row["email_text"])
        predicted, risk_level, score, flags = predict_label(email_text)

        total += 1
        level_counts[risk_level] = level_counts.get(risk_level, 0) + 1
        if actual == "phishing":
            counts["tp" if predicted == "phishing" else "fn"] += 1
        else:
            counts["fp" if predicted == "phishing" else "tn"] += 1

        if show_rows:
            print(f"{total}. actual={actual} predicted={predicted} "
                  f"risk={risk_level} score={score} flags={len(flags)}")

    tp, fp, tn, fn = counts["tp"], counts["fp"], counts["tn"], counts["fn"]
    actual_phishing = tp + fn
    actual_legitimate = fp + tn

    def ratio(numerator, denominator):
        return numerator / denominator if denominator else 0.0

    precision = ratio(tp, tp + fp)
    recall = ratio(tp, actual_phishing)

    return {
        "dataset": str(path),
        "total": total,
        "actual_phishing": actual_phishing,
        "actual_legitimate": actual_legitimate,
        "correct": tp + tn,
        "true_positive": tp,
        "true_negative": tn,
        "false_positive": fp,
        "false_negative": fn,
        "accuracy": ratio(tp + tn, total),
        # Rates are expressed per class, not per dataset: "1.3% of all emails
        # were false positives" hides how many legitimate emails there were to
        # get wrong in the first place.
        "false_positive_rate": ratio(fp, actual_legitimate),
        "false_negative_rate": ratio(fn, actual_phishing),
        "precision": precision,
        "recall": recall,
        "f1_score": ratio(2 * precision * recall, precision + recall),
        "risk_levels": level_counts,
    }


def print_summary(results):
    print("\nEvaluation summary")
    print(f"Dataset: {results['dataset']}")
    print(f"Total emails: {results['total']} "
          f"({results['actual_phishing']} phishing / {results['actual_legitimate']} legitimate)")
    print(f"Correct predictions: {results['correct']}")
    print(f"Accuracy: {results['accuracy']:.2%}")
    print(f"Precision: {results['precision']:.2%}")
    print(f"Recall: {results['recall']:.2%}")
    print(f"F1 score: {results['f1_score']:.2%}")
    print("\nConfusion matrix")
    print(f"  True positives  (phishing caught):      {results['true_positive']}")
    print(f"  False negatives (phishing missed):      {results['false_negative']} "
          f"({results['false_negative_rate']:.2%} of phishing)")
    print(f"  True negatives  (legitimate cleared):   {results['true_negative']}")
    print(f"  False positives (legitimate flagged):   {results['false_positive']} "
          f"({results['false_positive_rate']:.2%} of legitimate)")
    levels = results["risk_levels"]
    assigned = ", ".join(
        f"{name} {levels.get(name, 0)}" for name in ("Safe", "Suspicious", "Dangerous")
    )
    print(f"\nRisk levels assigned: {assigned}")


def main():
    parser = argparse.ArgumentParser(
        description="Evaluate PhishGuard against a labelled email dataset."
    )
    parser.add_argument("dataset_path", help="CSV/TSV/XLSX with `label` and `email_text` columns.")
    parser.add_argument("--quiet", action="store_true",
                        help="Only print the summary, not every row.")
    parser.add_argument("--json", metavar="PATH", help="Also write the results to a JSON file.")
    args = parser.parse_args()

    results = evaluate_dataset(args.dataset_path, show_rows=not args.quiet)
    print_summary(results)

    if args.json:
        Path(args.json).write_text(json.dumps(results, indent=2), encoding="utf-8")
        print(f"\nWrote {args.json}")


if __name__ == "__main__":
    main()
